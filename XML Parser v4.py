

from bs4 import BeautifulSoup
import pandas as pd
import spacy
from transformers import pipeline
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression
from sklearn.pipeline import make_pipeline
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Initialize NLP models
nlp = spacy.load("en_core_web_sm")  # spaCy model for NER
bert_classifier = pipeline("zero-shot-classification", model="facebook/bart-large-mnli")  # BERT for classification

# Example categories relevant to Lockheed Martin
categories = ["environmental", "safety", "health", "manufacturing", "military", "defense", "aviation", "space"]

# TF-IDF with Logistic Regression for impact evaluation
impact_model = make_pipeline(TfidfVectorizer(), LogisticRegression())
# Dummy training data for demonstration (replace with actual data)
train_texts = ["This regulation affects hazardous chemicals in aviation.", "New environmental standards for missile production."]
train_labels = ["High Impact", "Medium Impact"]
impact_model.fit(train_texts, train_labels)

def parse_regulation_info(xml_data):
    """Parse the XML data to extract and filter regulation information."""
    soup = BeautifulSoup(xml_data, 'xml')
    regulations = []

    for rin_info in soup.find_all('RIN_INFO'):
        rin = rin_info.find('RIN').text
        rule_title = rin_info.find('RULE_TITLE').text.lower()
        summary = rin_info.find('ABSTRACT').text.strip() if rin_info.find('ABSTRACT') else None
        agency = rin_info.find('AGENCY').find('NAME').text
        rule_stage = rin_info.find('RULE_STAGE').text

        # SpaCy NER processing
        doc = nlp(rule_title)
        entities = [ent.text for ent in doc.ents]

        # BERT-based classification
        bert_result = bert_classifier(rule_title, categories, multi_label=True)
        relevant_categories = [label for label, score in zip(bert_result['labels'], bert_result['scores']) if score > 0.5]

        # Initial impact evaluation using TF-IDF + Logistic Regression
        impact = impact_model.predict([rule_title])[0]

        # Add regulation if it's relevant based on BERT classification or NER entities
        if relevant_categories or any(entity in rule_title for entity in entities):
            regulations.append({
                'Agency': agency,
                'Agenda Stage of Rulemaking': rule_stage,
                'Title': rule_title.title(),
                'RIN': rin,
                'Summary': summary,
                'Initial Impact Evaluation': impact
            })

    return regulations

def save_to_excel(data, output_file):
    """Save extracted and filtered data to an Excel file, with color coding."""
    df = pd.DataFrame(data)
    df.to_excel(output_file, index=False)

    # Load the workbook to apply color coding
    wb = load_workbook(output_file)
    ws = wb.active

    # Define color fills for impact evaluation
    high_impact_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    medium_impact_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    low_impact_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    # Apply color coding based on the impact evaluation
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        impact = row[5].value  # Assuming "Initial Impact Evaluation" is the 6th column
        if impact == "High Impact":
            for cell in row:
                cell.fill = high_impact_fill
        elif impact == "Medium Impact":
            for cell in row:
                cell.fill = medium_impact_fill
        elif impact == "Low Impact":
            for cell in row:
                cell.fill = low_impact_fill

    # Save the updated workbook
    wb.save(output_file)

# Example usage:
if __name__ == "__main__":
    # Load your XML data (assuming it's in a string or file)
    with open('unified_agenda.xml', 'r') as file:
        xml_data = file.read()
    
    regulations = parse_regulation_info(xml_data)
    output_file = 'filtered_unified_agenda_regulations.xlsx'
    save_to_excel(regulations, output_file)
    print(f"Filtered data saved to {output_file}")
